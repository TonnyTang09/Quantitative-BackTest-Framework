import numpy as np
import pandas as pd
import math
from Data import Data
import Strategy
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from iFinDPy import *
from datetime import datetime, timedelta
import os


class Backtest:

    def __init__(self, underlying, strategy):
        """
        初始化Backtest对象 需要传入的参数包括
        underlying：一个Data类对象 代表当前回测的标的 如 Data('IZL.DCE')
        strategy：一个Strategy类对象（实际是继承了Strategy的某个子类） 代表当前回测的策略 如Strategy.momentum_1h()
        """

        # 初始化一些类变量
        # self.start_date = start_date  # 回测开始时间
        #
        # self.end_date = end_date  # 回测结束时间

        self.underlying = underlying  # 回测标的（以Data类对象传过来 因为计算指标的函数封装在Data类中）

        self.data = underlying.data  # 回测数据（通过相关类方法获取并计算相关指标后的数据）

        self.strategy = strategy  # 回测策略（以Strategy类对象传过来 因为开仓平仓条件封装在Strategy类中）

        self.trading_details = {
            'open': [],  # 记录每一笔开仓价格
            'close': [],  # 记录每一笔调仓价格
            'open time': [],  # 记录每一笔开仓时间
            'close time': [],  # 记录每一笔调仓时间
            'direction': [],  # 记录每一笔交易的方向 1是空 0是多
            'total profit': 0,  # 记录整个策略在回测区间的总收益
            'total fee': 0,  # 记录整个策略在回测区间的总交易费率
            'balance': [10000000],  # 记录账户的实际余额 只在每一次调仓后更新（无操作时不加入新的数值）
            'cumulative_return': [10000000],  # 记录账户实际收益 每天更新（无操作时会以前一个值填充，这一点跟balance区别）
            'pnl': [10000000],  # 记录账户浮盈浮亏 每天更新
            'size': 0  # 记录当前持仓大小
        }

        self.long_margin = self.underlying.long_margin  # 当前标的开多的保证金比例 因为每次计算时都要访问 所以通过类变量储存

        self.short_margin = self.underlying.short_margin  # 当前标的开空的保证金比例

        self.underlying_multiplier = self.underlying.underlying_multiplier  # 当前标的的合约乘数 如铁矿石是100

        self.transaction_rate = self.underlying.transaction_rate  # 当前标的的交易费率

        self.signal = 0  # 当前持仓信号 -1代表无持仓 0代表多 1代表空

        self.next_open_size = 0  # 移仓换月时需要保存一下当前持仓大小 因为trading details里的size会在换月时被调仓函数改成0 不方便访问

        self.next_open_direction = -1  # 移仓换月时开单方向（保持空或者保持多）

        self.force_open_signal = 0  # 控制是否该换月的变量 当出现主力合约更替时 更改为1 进行换月 之后再更改回0

        self.indicator_to_draw = []

    def open_position(self, temp_price, temp_date, i, open_size, add_up=False):
        """
        开仓函数 当策略出现开仓信号后调用 需要传入参数
        temp_price：当前价格
        temp_date：当前时间
        i：当前时间步在整个回测区间内的索引（第几个时间步）
        open_size：开仓大小 由策略的开仓信号函数产生

        直接对类变量进行修改 所以不需要返回值
        """
        if not add_up:
            # 检查当前是空开还是多开 相应在交易记录和数据中做标记 方便后续画图统计 用temp_margin记录当前交易需要的保证金比例
            if self.signal == -1:
                self.trading_details['direction'].append(-1)
                self.data.loc[i, 'direction'] = self.signal
                # self.data.loc[i, 'direction'] = self.signal
                self.data.loc[i, 'sign'] = '空开'
                # self.data.loc[i, 'sign'] = '空开'
                temp_margin = self.short_margin
            if self.signal == 1:
                self.trading_details['direction'].append(1)
                self.data.loc[i, 'direction'] = self.signal
                self.data.loc[i, 'sign'] = '多开'
                # self.data.loc[i, 'sign'] = '多开'
                temp_margin = self.long_margin

            # 记录当前开仓价格和开仓时间
            self.trading_details['open'].append(temp_price)
            self.trading_details['open time'].append(temp_date)

            # 如果当前时间步不处在换月 计算开仓大小
            if self.force_open_signal == 0:

                if open_size <= 1:
                    # 如果open_size小于等于1 代表用当前账户余额的比例来缴纳保证金 先计算能用多少资金 再计算具体开仓大小
                    temp_usable = self.trading_details['balance'][-1] * open_size

                if open_size > 1:
                    # 如果open_size大于1 代表当前开仓大小就是具体的多少手合约
                    temp_size = open_size
                else:
                    # 当前可用资金/（当前开仓价格*保证金比例）/合约乘数 = 可以开多少手合约 用floor来向下取整 再通过对10取余再乘10转换为整10手合约
                    temp_size = (math.floor(
                        temp_usable / (temp_price * temp_margin) / self.underlying_multiplier)) // 10 * 10
            else:
                # 如果当前时间步在换月节点 那么开仓大小保持为上一个时间步持仓大小
                temp_size = self.next_open_size

            # 将当前开仓大小分别记录在trading detials和数据中
            self.trading_details['size'] = temp_size
            self.data.loc[i, 'size'] = temp_size

        else:
            self.data.loc[i, 'sign'] = '加仓'
            if self.signal == -1:
                self.trading_details['direction'].append(-1)
                self.data.loc[i, 'direction'] = self.signal
                temp_margin = self.short_margin
            if self.signal == 1:
                self.trading_details['direction'].append(1)
                self.data.loc[i, 'direction'] = self.signal
                temp_margin = self.long_margin

            self.trading_details['open time'].append(temp_date)

            previous_size = self.data.loc[i-1, 'size']

            previous_cost = self.trading_details['open'][-1]

            if open_size > 1:
                # 如果open_size大于1 代表当前开仓大小就是具体的多少手合约
                temp_size = open_size
            else:
                previous_usable = previous_size * (previous_cost * temp_margin) * self.underlying_multiplier

                # 如果open_size小于等于1 代表用当前账户余额的比例来缴纳保证金 先计算能用多少资金 再计算具体开仓大小
                temp_usable = (self.trading_details['balance'][-1] - previous_usable) * open_size

                # 当前可用资金/（当前开仓价格*保证金比例）/合约乘数 = 可以开多少手合约 用floor来向下取整 再通过对10取余再乘10转换为整10手合约
                temp_size = (math.floor(
                    temp_usable / (temp_price * temp_margin) / self.underlying_multiplier)) // 10 * 10

            total_size = previous_size + temp_size

            temp_cost = previous_size / total_size * previous_cost + temp_size / total_size * temp_price

            # 记录当前开仓价格和开仓时间
            self.trading_details['open'].append(temp_cost)

            # 将当前开仓大小分别记录在trading detials和数据中
            self.trading_details['size'] = total_size
            self.data.loc[i, 'size'] = self.trading_details['size']

    def close_position(self, temp_price, temp_date, i, reduce_size):
        """
        调仓函数 当出现调仓信号后调用 需要传入的参数包括
        temp_price：当前价格
        temp_date：当前时间
        i：当前时间步在整个回测区间内的索引（第几个时间步）
        reduce_size：调仓大小 由策略的开仓信号函数产生

        直接对类变量进行修改 所以不需要返回值
        """

        # 记录一下当前调仓价格和调仓时间
        self.trading_details['close'].append(temp_price)
        self.trading_details['close time'].append(temp_date)

        # 当reduce_size为1时 说明全部平仓
        if reduce_size == 1:

            # 此时应该平仓的大小就是当前的全部持仓
            temp_sell_size = self.trading_details['size']  # Calculate the size of position to reduce now

            # 在数据中给当前时间步打上标记 方便画图
            if self.signal == -1:
                self.data.loc[i, 'sign'] = '空平'
            if self.signal == 1:
                self.data.loc[i, 'sign'] = '多平'

            # 确定全部平仓后 把持仓信号改为0
            self.signal = 0
        else:
            # 跟开仓函数中类似 计算当前调整的合约数
            if reduce_size < 1:
                temp_sell_size = math.floor(self.trading_details['size'] * reduce_size) // 10 * 10
            else:
                temp_sell_size = reduce_size

            self.data.loc[i, 'sign'] = '减平'

        # 当前调整合约数*（当前价格-开仓价格）*合约乘数 计算当前调仓后的收益/损失
        tmep_amount = (temp_sell_size * (temp_price - self.trading_details['open'][-1])) * self.underlying_multiplier

        # 当前调整合约数*（当前价格+开仓价格）*合约乘数*交易费率 计算当前调仓的手续费 因为开仓时没计算手续费 所以要加上开仓那部分的
        temp_fee = temp_sell_size * (
                temp_price + self.trading_details['open'][-1]) * self.underlying_multiplier * self.transaction_rate

        # 保留两位小数
        temp_fee = round(temp_fee, 2)

        # 将交易费累加到trading details中
        self.trading_details['total fee'] += temp_fee

        # 获取一下当前交易是空还是多（会影响收益/损失的计算）
        tmep_direction = self.trading_details['direction'][-1]

        # 获取一下当前账户余额 用来计算加上收益/损失后的账户余额
        temp_balance = self.trading_details['balance'][-1]  # Obtain the current balance

        # 当做多时 当前交易的收益/损失就是tmep_amount - temp_fee，账户余额变成temp_balance + tmep_amount
        # 当做空时 当前交易的收益/损失就是- tmep_amount - temp_fee，账户余额变成temp_balance - tmep_amount
        # 例子：100块时开仓 200块时平仓 tmep_amount的计算逻辑会返回 200 - 100 = 100
        # 如果是多头 收益是100 如果是空头 损失是100 这样可以理解下面的计算逻辑
        if tmep_direction == 1:
            temp_result = temp_balance + tmep_amount
            temp_profit = tmep_amount - temp_fee
        elif tmep_direction == -1:
            temp_result = temp_balance - tmep_amount
            temp_profit = - tmep_amount - temp_fee

        # 更改一下调仓后的持仓大小 如果reduce size是1 代表全平 要调整为0
        if reduce_size != 1:
            self.trading_details['size'] = self.trading_details['size'] - temp_sell_size
        else:
            self.trading_details['size'] = 0

        self.data.loc[i, 'size'] = self.trading_details['size']

        # 将收益累加到trading detials中
        self.trading_details['total profit'] += temp_profit

        # 将扣除手续费后的账户余额、浮盈浮亏、累积收益更新
        self.trading_details['balance'].append(temp_result - temp_fee)

        self.trading_details['cumulative_return'].append(temp_result - temp_fee)

        self.trading_details['pnl'].append(temp_result - temp_fee)

        self.data.loc[i, 'pnl'] = (temp_result - temp_fee)

    def check_transfer(self, i):
        """
        换月函数 每个时间步都要检查一下
        i：当前时间步在整个回测区间内的索引（第几个时间步）
        当检查符合换月时（主力合约更替） 返回True 代表需要换月了
        实现逻辑如下 先获取主连的当前交易合约和对应的时间序列 这样只需要检查当前交易合约和下一天交易合约是否一样 就可以知道应该在哪一天换月
        假设3月28号换月到i2409 那么在3月27晚上9点开盘后的第一分钟结束时 以i2405的收盘价平当前仓位 若此时下一合约上有开仓信号
        则以i2409的收盘价重开大小方向一致的仓位
        """

        # 获取当前时间步的主力合约代码和下一个时间步的主力合约代码
        temp_dominant = self.data.loc[i, '证券代码']

        next_dominant = self.data.loc[i+1, '证券代码']

        # 如果这两个主力合约代码不一样且当前有持仓 那么需要换月
        if temp_dominant != next_dominant and self.signal != 0:
            # 换月的逻辑是先平再开 记录当前平仓价格和时间
            temp_price = self.data.loc[i, 'close']

            temp_date = self.data.loc[i, 'date']

            # 保存当前仓位大小和方向给下一时间步开仓用
            self.next_open_size = self.trading_details['size']

            self.next_open_direction = self.signal

            self.force_open_signal = 1

            # 调用调仓函数 reduce size设置为1 全部平掉
            self.close_position(temp_price, temp_date, i, 1)

            return True

    # pnl记录浮盈浮亏（每个时间步动态变化） cumulative_return记录实际收益（也就是只有调仓的时候才会发生变化）拆开很多部分的原因在于解决不同
    # 场景。1、刚开仓/无仓位 用keep_pnl pnl和实际收益都跟上一个时间步相同；2、有仓位无动作 用adjust_pnl pnl需要变 实际收益不变；3、有仓位且
    # 有动作 这个计算放在adjust_position中进行
    def keep_pnl(self, i):
        """
        用来拷贝一下前一期浮盈浮亏和累积收益的函数 用在无仓位的情况（不管当前时间步是否新开 因为当前时间步新开仓位 到下一时间步才会计算浮盈浮亏）
        """
        temp_net = self.trading_details['pnl'][-1]

        self.trading_details['pnl'].append(temp_net)

        self.data.loc[i, 'pnl'] = temp_net

        temp_cumulative_return = self.trading_details['cumulative_return'][-1]

        self.trading_details['cumulative_return'].append(temp_cumulative_return)

    def adjust_pnl(self, temp_price, i):
        """
        用来更新一下浮盈浮亏 并拷贝前一时间步的 累积收益 用在有仓位无操作的情况 只需要更新浮盈浮亏 保持累积收益不变
        """

        # 这些计算跟adjust position中的计算一致
        temp_pnl = (temp_price - self.trading_details['open'][-1]) * self.trading_details[
            'size'] * self.underlying_multiplier

        temp_balance = self.trading_details['balance'][-1]

        if self.signal == 1:
            temp_result = temp_balance + temp_pnl

        else:
            temp_result = temp_balance - temp_pnl

        self.trading_details['pnl'].append(temp_result)

        self.data.loc[i, 'pnl'] = temp_result

        temp_cumulative_return = self.trading_details['cumulative_return'][-1]

        self.trading_details['cumulative_return'].append(temp_cumulative_return)

    def run_backtest(self):

        # 用data来代表类变量中的data 简洁一些
        data = self.data

        # 向data中新添加两列 一列sign代表当前持仓是多还是空（方便后续给excel格子上色） size代表持仓大小
        data['sign'] = np.nan

        data['size'] = np.nan

        data['pnl'] = np.nan

        data['direction'] = np.nan

        # 把收盘价转换成列表 方便遍历 在遍历时通过索引访问数据框的某个位置 这里把任意一列转换都可以 无所谓 因为长度是一样的
        close = data['close'].tolist()

        # 边界情况 要把第一个时间戳放进来 跟初始状态对齐 因为在画出收益曲线时 第一个点一定要是第一个时间戳和初始账户余额
        self.trading_details['close time'].append(data.loc[0, 'date'])

        data.loc[0, 'pnl'] = self.trading_details['balance'][0]

        # 开始遍历每一个时间步
        for i in range(len(close)):

            # 获取当前收盘价和时间
            temp_price = close[i]

            temp_date = data.loc[i, 'date']

            # 先检查是否换月了 因为有可能出现下一个主力合约并没有触发开仓条件 但是必须要把仓位换过来
            if self.force_open_signal == 1:

                # 把控制换月的信号改掉 防止错误进入
                self.force_open_signal = 0

                if self.strategy.open_criterion(data, i) != 0:
                    # 将当前持仓方向由换月上一时间步保存的替换
                    self.signal = self.next_open_direction

                    # 调用开仓函数 仓位大小保持不变
                    self.open_position(temp_price, temp_date, i, self.next_open_size)

                    # 新开仓位 浮盈/浮亏和累积收益不变 所以调用keep pnl
                    self.keep_pnl(i)

                    # 每一个时间步只允许一个动作 或者 无动作 所以要continue到下一个时间步
                    continue

            # 如果当前没有持有头寸 检查是否满足开仓条件
            if self.signal == 0:
                # 调用策略的检查开仓条件函数 open direction记录开仓方向 open size记录开仓大小
                open_size = self.strategy.open_criterion(data, i)

                # 如果open size是0 那么说明不满足开仓条件
                if open_size == 0:

                    # 向data中填充相应指示 保留当前持仓不变
                    self.data.loc[i, 'sign'] = '无操作'

                    self.data.loc[i, 'size'] = self.trading_details['size']

                    self.trading_details['direction'].append(0)

                    self.data.loc[i, 'direction'] = 0

                    # 多加一个判断条件 防止第一个时间步不开仓时引起索引出界
                    if i >= 1:
                        # 无操作 要调用keep pnl
                        self.keep_pnl(i)

                    continue

                # 如果open size不是0 代表可以开仓
                if open_size > 0:
                    self.signal = 1
                else:
                    self.signal = -1

                # 调用开仓函数 开仓大小由上面策略的开仓信号函数产生
                self.open_position(temp_price, temp_date, i, abs(open_size))

                if i >= 1:
                    # 新开仓 调用keep pnl
                    self.keep_pnl(i)

                continue

            # 当有持仓时 更新一下trading details中的持仓方向
            self.trading_details['direction'].append(self.signal)

            self.data.loc[i, 'direction'] = self.signal

            # 检查是否换月 如果需要换月 那就不需要再看后面的条件 直接强平
            if i < len(close) - 1:

                if self.check_transfer(i):
                    continue

            # 如果当前时间步是最后一个时间步了 强平
            if i == len(close) - 1:
                self.close_position(temp_price, temp_date, i, 1)

                break

            # 如果当前有持仓 检查调仓条件 可能是减仓加仓或平仓（加仓的逻辑暂时还没写好 只支持减仓）
            if self.signal != 0:
                # 返回 0 不操作 其他调用 adjust_positon
                temp_reduce_size = self.strategy.close_criterion(data, i, self.signal)

                if temp_reduce_size != 0:

                    keep_signal = self.signal

                    self.close_position(temp_price, temp_date, i, temp_reduce_size)

                    temp_reverse_size = self.strategy.open_criterion(data, i)

                    if temp_reverse_size * keep_signal < 0:

                        if temp_reverse_size > 0:
                            self.signal = 1
                        else:
                            self.signal = -1

                        self.open_position(temp_price, temp_date, i, abs(temp_reverse_size))

                        if temp_reverse_size > 0:
                            self.data.loc[i, 'sign'] = '空平+多开'
                        else:
                            self.data.loc[i, 'sign'] = '多平+空开'

                        self.keep_pnl(i)

                    continue

                temp_add_size = self.strategy.open_criterion(data, i, self.signal)

                if temp_add_size != 0:
                    self.open_position(temp_price, temp_date, i, temp_add_size, add_up=True)

                    self.adjust_pnl(temp_price, i)

                    continue

            # 如果运行到这里表明当前有仓位但不触发调仓信号
            self.data.loc[i, 'sign'] = '无操作'

            self.data.loc[i, 'size'] = self.trading_details['size']

            if i >= 1:
                # 要计算浮盈浮亏 所以调用adjust pnl
                self.adjust_pnl(temp_price, i)

        # 此时已经遍历完了整个数据 完成了回测 下面是把记录好的数据放到数据框中 并把想要的数据保存到excel中
        self.data['pnl'] = self.data['pnl'].apply(lambda x: int(x))

        self.data['pnl'] = self.data['pnl'].apply(lambda x: "{:,}".format(x))

        self.trading_details['total profit'] = round(self.trading_details['total profit'], 2)

        self.trading_details['total fee'] = round(self.trading_details['total fee'], 2)

        indicator_to_draw = []

        for key in self.strategy.indicator_require.keys():

            for ele in self.strategy.indicator_require[key]:

                if isinstance(ele, bool):
                    temp_indicator = str(key)
                else:
                    temp_indicator = str(key) + str(ele)

                indicator_to_draw.append(temp_indicator)

        selected_columns = ['证券代码', 'date', 'open', 'high', 'low', 'close', 'change_ratio'] + indicator_to_draw + [
            'sign', 'size', 'direction',
            'pnl']  # 'ma60_daily',

        selected_data = self.data[selected_columns]

        column_index = selected_data.columns.get_loc('direction')

        color_index = selected_data.columns.get_loc('sign')

        selected_data['change_ratio'] = round(selected_data['change_ratio'], 2) / 100

        selected_data['change_ratio'] = np.array(["{:.2%}".format(num) for num in selected_data['change_ratio']])

        selected_data = selected_data.rename(columns={'证券代码': '合约代码',
                                                      'date': '日期时间',
                                                      'open': '开盘价',
                                                      'high': '最高价',
                                                      'low': '最低价',
                                                      'close': '收盘价',
                                                      'change_ratio': '涨跌幅',
                                                      'size': '现有仓位大小',
                                                      'sign': '操作',
                                                      'direction': '头寸',
                                                      'pnl': '权益',
                                                      'cumulative_return': '累积收益'
                                                      })

        selected_data['日期时间'] = pd.to_datetime(selected_data['日期时间'])

        self.data = selected_data

        # 创建一个新的excel工作表
        wb = Workbook()
        ws = wb.active

        # 写入列名和数据
        ws.append(selected_data.columns.tolist())

        for row in selected_data.itertuples(index=False):
            ws.append(row)

        # 给持有多单的时间步填充红色 空单的时间步填充绿色
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            direction = row[column_index].value
            if direction == 1:
                row[color_index].fill = PatternFill('solid', start_color='00FF0000')
                continue
            elif direction == -1:
                row[color_index].fill = PatternFill('solid', start_color='0000FF00')
                continue

        # 获取当前文件所在的目录路径
        current_directory = os.getcwd()

        # 创建 "raw data" 文件夹路径
        trading_details_directory = os.path.join(current_directory, "Trading Details")

        # 确保 "raw data" 文件夹存在，如果不存在则创建
        if not os.path.exists(trading_details_directory):
            os.makedirs(trading_details_directory)

        # 保存excel文件到指定文件夹
        output_file_path = os.path.join(trading_details_directory,
                                        f'{self.strategy.name}_for_{self.underlying.code}.xlsx')
        wb.save(output_file_path)
